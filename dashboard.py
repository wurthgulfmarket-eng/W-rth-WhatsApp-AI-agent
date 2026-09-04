"""
Internal analytics dashboard: message/customer stats, a read-only transcript
viewer, and an Excel export - all reading from the same Postgres database the
webhook writes to (storage/store.py).

Protected by a username/password login (DASHBOARD_ADMIN_USERNAME /
DASHBOARD_ADMIN_PASSWORD env vars), backed by a signed session cookie
(itsdangerous) rather than a server-side session store, so login survives
across the app's multiple worker/background threads without extra state.
"""
import io
from datetime import date, datetime, timedelta, timezone

from fastapi import APIRouter, Form, Request, Response
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from itsdangerous import BadSignature, SignatureExpired, URLSafeTimedSerializer

from config import config
from storage import store

router = APIRouter()

SESSION_COOKIE_NAME = "wurth_dashboard_session"
SESSION_MAX_AGE_SECONDS = 60 * 60 * 12  # 12 hours

LOGO_URL = (
    "https://eshop.wurth.ae/is-bin/intershop.static/WFS/3890-B1-Site/-/en_US/"
    "webkit_bootstrap/dist/img/wuerth-logo.svg"
)


def _serializer():
    return URLSafeTimedSerializer(config.DASHBOARD_SESSION_SECRET, salt="dashboard-session")


def _is_logged_in(request: Request) -> bool:
    cookie = request.cookies.get(SESSION_COOKIE_NAME)
    if not cookie:
        return False
    try:
        data = _serializer().loads(cookie, max_age=SESSION_MAX_AGE_SECONDS)
        return data.get("user") == config.DASHBOARD_ADMIN_USERNAME
    except (BadSignature, SignatureExpired):
        return False


def _logged_in_user(request: Request) -> str | None:
    cookie = request.cookies.get(SESSION_COOKIE_NAME)
    if not cookie:
        return None
    try:
        return _serializer().loads(cookie, max_age=SESSION_MAX_AGE_SECONDS).get("user")
    except (BadSignature, SignatureExpired):
        return None


def _default_date_range():
    end = date.today()
    start = end - timedelta(days=30)
    return start.isoformat(), end.isoformat()


@router.get("/dashboard/login", response_class=HTMLResponse)
def login_page(error: str = ""):
    return HTMLResponse(_render_login_html(error))


@router.post("/dashboard/login")
def login_submit(username: str = Form(""), password: str = Form("")):
    if not config.DASHBOARD_ADMIN_USERNAME or not config.DASHBOARD_ADMIN_PASSWORD:
        return HTMLResponse(
            _render_login_html("Dashboard login is not configured yet - set DASHBOARD_ADMIN_USERNAME "
                                "and DASHBOARD_ADMIN_PASSWORD in the app's environment variables."),
            status_code=500,
        )

    if username == config.DASHBOARD_ADMIN_USERNAME and password == config.DASHBOARD_ADMIN_PASSWORD:
        token = _serializer().dumps({"user": username})
        resp = RedirectResponse(url="/dashboard", status_code=303)
        resp.set_cookie(
            SESSION_COOKIE_NAME, token,
            max_age=SESSION_MAX_AGE_SECONDS, httponly=True, samesite="lax",
        )
        return resp

    return HTMLResponse(_render_login_html("Incorrect username or password."), status_code=401)


@router.get("/dashboard/logout")
def logout():
    resp = RedirectResponse(url="/dashboard/login", status_code=303)
    resp.delete_cookie(SESSION_COOKIE_NAME)
    return resp


@router.post("/dashboard/leads/{lead_id}/mark-false-positive")
def mark_false_positive(lead_id: int, request: Request, phone: str = Form(""), enquiry_text: str = Form(""),
                         start: str = Form(""), end: str = Form("")):
    if not _is_logged_in(request):
        return RedirectResponse(url="/dashboard/login", status_code=303)
    store.mark_lead_false_positive(lead_id, phone, enquiry_text, marked_by=_logged_in_user(request))
    return RedirectResponse(url=f"/dashboard?start={start}&end={end}", status_code=303)


@router.post("/dashboard/conversations/{conversation_id}/manual-escalate")
def manual_escalate(conversation_id: int, request: Request, start: str = Form(""), end: str = Form(""),
                     phone: str = Form(""), page: int = Form(1)):
    """Manually escalates one inbound message the automated pipeline
    didn't flag as a lead - for enquiries the AI genuinely can't verify on
    its own (e.g. an image with no caption, an ambiguous one-word reply)
    but a human reviewing the transcript can tell is a real lead. Reuses
    the same lead + notification path as an automated escalation, so it
    shows up on the Leads panel and sends the same rep/ops alert."""
    if not _is_logged_in(request):
        return RedirectResponse(url="/dashboard/login", status_code=303)

    # Local import avoids a circular import - main.py imports dashboard.py
    # to mount this router, so dashboard.py can't import main.py at module
    # load time.
    from main import _notify_escalation

    msg = store.get_conversation_message(conversation_id)
    if msg and msg["direction"] == "in" and not msg["escalated"]:
        customer = store.get_customer(msg["phone"])
        store.mark_conversation_escalated(conversation_id)
        store.get_or_open_lead(msg["phone"], conversation_id, priority="medium")
        _notify_escalation(conversation_id, msg["phone"], msg["message"], customer)

    return RedirectResponse(url=f"/dashboard?start={start}&end={end}&phone={phone}&page={page}", status_code=303)


_VALID_OUTCOMES = {"new", "contacted", "quoted", "won", "lost"}


@router.post("/dashboard/leads/{lead_id}/set-outcome")
def set_outcome(lead_id: int, request: Request, outcome: str = Form(""), amount: str = Form(""),
                 note: str = Form(""), start: str = Form(""), end: str = Form(""), lpage: int = Form(1)):
    """Lets an admin log what happened with a lead after it was sent to a
    rep (New/Contacted/Quoted/Won/Lost) - the dashboard-side counterpart to
    the WhatsApp-keyword shortcut a rep can also use (see
    ai.agent.try_extract_rep_outcome_signal). Rejects an outcome value
    outside the known set rather than trusting the form blindly."""
    if not _is_logged_in(request):
        return RedirectResponse(url="/dashboard/login", status_code=303)

    if outcome in _VALID_OUTCOMES:
        parsed_amount = None
        if outcome == "won" and amount.strip():
            try:
                parsed_amount = float(amount.strip().replace(",", ""))
            except ValueError:
                parsed_amount = None
        store.set_lead_outcome(lead_id, outcome, updated_by=_logged_in_user(request),
                                amount=parsed_amount, note=note.strip() or None)

    return RedirectResponse(url=f"/dashboard?start={start}&end={end}&lpage={lpage}", status_code=303)


_TRANSCRIPT_PAGE_SIZE = 20
_CUSTOMERS_PAGE_SIZE = 20
_LEADS_LIST_PAGE_SIZE = 20
_LEADS_SUMMARY_PAGE_SIZE = 20


@router.get("/dashboard", response_class=HTMLResponse)
def dashboard_page(request: Request, start: str = "", end: str = "", phone: str = "", rep_phone: str = "",
                    page: int = 1, cpage: int = 1, lpage: int = 1, spage: int = 1):
    if not _is_logged_in(request):
        return RedirectResponse(url="/dashboard/login", status_code=303)

    default_start, default_end = _default_date_range()
    start = start or default_start
    end = end or default_end
    page = max(page, 1)
    cpage = max(cpage, 1)
    lpage = max(lpage, 1)
    spage = max(spage, 1)

    stats = store.get_stats(start, end)
    daily = store.get_daily_counts(start, end)
    outcome_breakdown = store.get_outcome_breakdown(start, end)
    customers, customers_total = store.get_customers_summary(start, end, page=cpage, page_size=_CUSTOMERS_PAGE_SIZE)
    leads_summary, leads_summary_total = store.get_leads_summary(start, end, page=spage, page_size=_LEADS_SUMMARY_PAGE_SIZE)
    # Unpaginated top-10 for the bar chart - a separate call from the paginated
    # table below it, since the chart always shows the leaderboard regardless
    # of which table page is selected.
    leads_summary_all, _ = store.get_leads_summary(start, end)
    leads_list, leads_list_total = store.get_leads_list(start, end, page=lpage, page_size=_LEADS_LIST_PAGE_SIZE)
    rep_replies = store.get_rep_replies_list(start, end)
    reps = store.get_reps_summary(start, end)
    if phone:
        transcript, transcript_total = store.get_conversation(phone, start, end, page=page, page_size=_TRANSCRIPT_PAGE_SIZE)
    else:
        transcript, transcript_total = None, 0
    rep_transcript = store.get_rep_transcript(rep_phone, start, end) if rep_phone else None

    return HTMLResponse(_render_dashboard_html(
        start, end, stats, daily, outcome_breakdown, customers, customers_total, cpage,
        leads_summary, leads_summary_total, spage, leads_summary_all, leads_list, leads_list_total, lpage,
        rep_replies, reps, phone, transcript, transcript_total, page, rep_phone, rep_transcript,
    ))


@router.get("/dashboard/export")
def export_excel(request: Request, start: str = "", end: str = ""):
    if not _is_logged_in(request):
        return Response("Forbidden - please log in", status_code=403)

    default_start, default_end = _default_date_range()
    start = start or default_start
    end = end or default_end

    from openpyxl import Workbook

    wb = Workbook()

    ws = wb.active
    ws.title = "Messages"
    ws.append(["Timestamp (UTC)", "Phone", "Company", "Direction", "Message", "Escalated"])
    for row in store.get_all_messages(start, end):
        ws.append([
            str(row["created_at"]), row["phone"], row["company_name"],
            "Customer" if row["direction"] == "in" else "Bot",
            row["message"], "Yes" if row["escalated"] else "No",
        ])
    for col_letter, width in zip("ABCDEF", [26, 16, 24, 10, 60, 10]):
        ws.column_dimensions[col_letter].width = width

    ws2 = wb.create_sheet("Customers")
    ws2.append(["Phone", "Company", "Sales Rep", "Message Count", "Last Message (UTC)"])
    all_customers, _ = store.get_customers_summary(start, end)  # page_size=None -> every customer, not just one dashboard page
    for row in all_customers:
        ws2.append([row["phone"], row["company_name"], row["rep_name"], row["message_count"], str(row["last_message_at"])])
    for col_letter, width in zip("ABCDE", [16, 24, 20, 14, 26]):
        ws2.column_dimensions[col_letter].width = width

    ws3 = wb.create_sheet("Daily Summary")
    ws3.append(["Date", "Messages Received", "Replies Sent"])
    for row in store.get_daily_counts(start, end):
        ws3.append([row["day"], row["received"], row["sent"]])
    for col_letter, width in zip("ABC", [14, 18, 14]):
        ws3.column_dimensions[col_letter].width = width

    ws4 = wb.create_sheet("Sales Leads")
    ws4.append(["Sales Rep", "Leads Generated", "Unique Customers", "Last Lead (UTC)", "Failed Notifications"])
    all_leads_summary, _ = store.get_leads_summary(start, end)  # page_size=None -> every rep
    for row in all_leads_summary:
        ws4.append([row["rep_name"], row["lead_count"], row["customer_count"], str(row["last_lead_at"]), row["failed_notifications"]])
    for col_letter, width in zip("ABCDE", [22, 16, 18, 26, 18]):
        ws4.column_dimensions[col_letter].width = width

    ws5 = wb.create_sheet("Lead Details")
    ws5.append(["Timestamp (UTC)", "Phone", "Company", "Sales Rep", "Customer Enquiry", "Priority", "Status", "Delivery", "Rep Response", "Response Confidence", "Marked False Positive"])
    all_leads_list, _ = store.get_leads_list(start, end)  # page_size=None -> every lead
    for row in all_leads_list:
        response_confidence = {"context_match": "Confirmed", "fallback_most_recent": "Best guess"}.get(row.get("rep_reply_method"), "")
        ws5.append([
            str(row["created_at"]), row["phone"], row["company_name"], row["rep_name"], row["enquiry_text"],
            (row.get("priority") or "").title(), row["status"], row["delivery_status"],
            row.get("rep_reply_text") or "", response_confidence,
            "Yes" if row.get("false_positive") else "No",
        ])
    for col_letter, width in zip("ABCDEFGHIJK", [26, 16, 24, 20, 60, 10, 10, 14, 40, 16, 16]):
        ws5.column_dimensions[col_letter].width = width

    ws6 = wb.create_sheet("Rep Replies")
    ws6.append(["Timestamp (UTC)", "Rep", "Rep Phone", "Customer", "Customer Phone", "Reply", "Match Confidence"])
    _CONFIDENCE_LABELS_XLSX = {"context_match": "Confirmed", "fallback_most_recent": "Best guess", "unresolved": "Unresolved"}
    for row in store.get_rep_replies_list(start, end):
        ws6.append([
            str(row["created_at"]), row["rep_name"], row["rep_phone"], row["company_name"], row["customer_phone"],
            row["reply_text"], _CONFIDENCE_LABELS_XLSX.get(row["resolution_method"], row["resolution_method"]),
        ])
    for col_letter, width in zip("ABCDEFG", [26, 20, 16, 24, 16, 60, 16]):
        ws6.column_dimensions[col_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"wurth-whatsapp-report_{start}_to_{end}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _esc(s) -> str:
    return str(s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


# All timestamps are stored in Postgres as UTC (see storage/store.py's
# datetime.now(timezone.utc) calls) - the dashboard is used by staff in the
# UAE, so every displayed time is converted to Gulf Standard Time (UTC+4,
# no daylight saving) here rather than showing raw UTC clock time.
_DISPLAY_TZ = timezone(timedelta(hours=4))


def _fmt_ts(ts) -> str:
    if isinstance(ts, datetime):
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
        return ts.astimezone(_DISPLAY_TZ).strftime("%Y-%m-%d %H:%M")
    return str(ts)[:16].replace("T", " ")


_DELIVERY_PILL_LABELS = {"delivered": "Delivered", "failed": "Failed", "pending": "Pending"}
_STATUS_PILL_LABELS = {"open": "Open", "closed": "Closed"}
_PRIORITY_PILL_LABELS = {"high": "High", "medium": "Medium", "low": "Low"}


def _delivery_pill(status: str, summary: str) -> str:
    label = _DELIVERY_PILL_LABELS.get(status, status)
    title = f' title="{_esc(summary)}"' if summary else ""
    return f'<span class="pill {_esc(status)}"{title}>{label}</span>'


def _status_pill(status: str) -> str:
    label = _STATUS_PILL_LABELS.get(status, status)
    return f'<span class="pill status-{_esc(status)}">{label}</span>'


def _priority_pill(priority: str | None) -> str:
    if not priority:
        return '<span class="muted">-</span>'
    label = _PRIORITY_PILL_LABELS.get(priority, priority)
    return f'<span class="pill priority-{_esc(priority)}">{label}</span>'


def _false_positive_pill() -> str:
    return '<span class="pill fp">Marked not-a-lead</span>'


_OUTCOME_PILL_LABELS = {"new": "New", "contacted": "Contacted", "quoted": "Quoted", "won": "Won", "lost": "Lost"}


def _outcome_pill(outcome: str, amount) -> str:
    label = _OUTCOME_PILL_LABELS.get(outcome, outcome or "New")
    if outcome == "won" and amount:
        label += f" (AED {amount:,.0f})"
    return f'<span class="pill outcome-{_esc(outcome or "new")}">{_esc(label)}</span>'


def _rep_reply_cell(reply_text: str | None, reply_at, method: str | None) -> str:
    if not reply_text:
        return '<span class="muted">No reply yet</span>'
    guess_badge = ' <span class="pill guess" title="Best guess - the rep did not reply directly to the alert, so this is their most recent open lead, not a confirmed match">Best guess</span>' if method == "fallback_most_recent" else ""
    truncated = reply_text if len(reply_text) <= 60 else reply_text[:57] + "..."
    return f'<span title="{_esc(reply_text)}">{_esc(truncated)}</span> <span class="muted">{_fmt_ts(reply_at)}</span>{guess_badge}'


_BASE_STYLE = """
  * { box-sizing: border-box; }
  body { font-family: -apple-system, Segoe UI, Arial, sans-serif; margin: 0; background: #f7f7f9; color: #1a1a1a; }
  header { background: #ffffff; color: #1a1a1a; padding: 14px 24px; display: flex; align-items: center; gap: 14px;
           flex-wrap: wrap; border-bottom: 1px solid #eceef1; }
  header img.logo { height: 26px; }
  header h1 { margin: 0; font-size: 1.05em; font-weight: 700; flex: 1; min-width: 0; color: #1a1a1a; }
  header a.logout { color: #6b7280; text-decoration: none; font-size: 0.85em; white-space: nowrap; font-weight: 600; }
  header a.logout:hover { color: #CC0000; }
"""


def _render_login_html(error: str = "") -> str:
    error_html = f'<p class="error">{_esc(error)}</p>' if error else ""
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Würth WhatsApp Agent - Login</title>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta name="robots" content="noindex, nofollow">
<style>
  * {{ box-sizing: border-box; }}
  body {{ font-family: -apple-system, Segoe UI, Arial, sans-serif; margin: 0; background: #f5f6f8; color: #1a1a1a;
         display: flex; align-items: center; justify-content: center; min-height: 100vh; padding: 16px; }}
  .card {{ background: white; border-radius: 10px; padding: 32px 28px; width: 100%; max-width: 360px; box-shadow: 0 2px 12px rgba(0,0,0,0.08); }}
  .card img {{ height: 32px; display: block; margin: 0 auto 20px auto; }}
  h1 {{ font-size: 1.1em; text-align: center; margin: 0 0 20px 0; }}
  label {{ display: block; font-size: 0.85em; color: #555; margin-bottom: 4px; margin-top: 14px; }}
  input {{ width: 100%; padding: 10px 12px; border: 1px solid #ccc; border-radius: 6px; font-size: 1em; }}
  button {{ width: 100%; margin-top: 20px; background: #CC0000; color: white; border: none; padding: 11px; border-radius: 6px; font-size: 1em; cursor: pointer; }}
  .error {{ color: #CC0000; font-size: 0.85em; margin-top: 12px; text-align: center; }}
</style>
</head>
<body>
  <form class="card" method="post" action="/dashboard/login">
    <img src="{LOGO_URL}" alt="Würth">
    <h1>WhatsApp Agent Dashboard</h1>
    <label>Username</label>
    <input type="text" name="username" autocomplete="username" required autofocus>
    <label>Password</label>
    <input type="password" name="password" autocomplete="current-password" required>
    <button type="submit">Log in</button>
    {error_html}
  </form>
</body>
</html>"""


def _render_pagination(base_params: dict, page_param: str, current_page: int, total_pages: int) -> str:
    """base_params are the OTHER query params to preserve on each page link
    (start/end and, for the transcript, which phone is selected) - page_param
    is which page-number param this particular pagination control updates
    (e.g. "page" for the customer transcript, "cpage" for the customers
    list), so the two pagination controls on the same page don't collide."""
    if total_pages <= 1:
        return ""
    query_prefix = "&".join(f"{k}={v}" for k, v in base_params.items())
    links = []
    for p in range(1, total_pages + 1):
        if p == current_page:
            links.append(f'<span class="page-num active">{p}</span>')
        else:
            links.append(f'<a class="page-num" href="?{query_prefix}&{page_param}={p}">{p}</a>')
    return f'<div class="pagination">{"".join(links)}</div>'


# ===== Chart building blocks =====
# All charts are hand-built inline SVG (no external chart library, keeping
# this app dependency-free) following the dataviz skill's method: form
# picked by the data's job, color assigned last, categorical hues from a
# validated palette (see the module-level comment above _OUTCOME_COLORS),
# thin marks, a surface gap between touching marks, direct labels used
# sparingly, and a legend whenever 2+ series are on screen.

# Brand accent - Wurth's red, used as the sole "identity" hue everywhere a
# single series/accent is enough (line-chart "Sent" series, bar chart,
# buttons, focus states). Never placed adjacent to green in the same chart
# (see _OUTCOME_COLORS below) - that pairing fails CVD separation, so the
# outcome donut uses a validated near-red instead for its "Lost" slice.
_BRAND_RED = "#CC0000"
_INK = "#1a1a1a"
_INK_MUTED = "#6b7280"
_GRID = "#eef0f3"

# Validated categorical palette for the 5-stage lead-outcome donut (see
# scripts/validate_palette.js "#2a78d6,#1baf7a,#eda100,#008300,#e34948" -
# ALL CHECKS PASS in light mode). Two slots (Contacted, Quoted) sit under
# the 3:1 contrast floor, so those slices always ship a direct label with
# their count (the "relief rule") rather than relying on hue alone.
_OUTCOME_COLORS = {
    "new": "#2a78d6",
    "contacted": "#1baf7a",
    "quoted": "#eda100",
    "won": "#008300",
    "lost": "#e34948",
}
# Two-series line chart palette (Received vs Sent) - validated separately
# (ALL CHECKS PASS): blue reads as "inbound/neutral", brand red as "our
# reply" - an intentional, not arbitrary, pairing of identity to meaning.
_LINE_COLOR_RECEIVED = "#2a78d6"
_LINE_COLOR_SENT = _BRAND_RED


def _sparkline_svg(values: list, color: str, width: int = 100, height: int = 32) -> str:
    """A tiny trend line for a stat tile - no axes, no gridlines, just the
    shape (per the dataviz skill's stat-tile contract: value + delta +
    12-point sparkline). Renders nothing (caller shows a flat dash instead)
    if there isn't enough data to draw a trend."""
    if len(values) < 2:
        return ""
    lo, hi = min(values), max(values)
    span = (hi - lo) or 1
    pad = 3
    step = (width - 2 * pad) / (len(values) - 1)
    points = [
        (pad + i * step, height - pad - ((v - lo) / span) * (height - 2 * pad))
        for i, v in enumerate(values)
    ]
    path = "M " + " L ".join(f"{x:.1f},{y:.1f}" for x, y in points)
    last_x, last_y = points[-1]
    return f"""<svg viewBox="0 0 {width} {height}" width="{width}" height="{height}" class="sparkline">
        <path d="{path}" fill="none" stroke="{color}" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
        <circle cx="{last_x:.1f}" cy="{last_y:.1f}" r="3" fill="{color}" stroke="#fff" stroke-width="1.5"/>
    </svg>"""


def _line_chart_svg(daily: list, width: int = 640, height: int = 220) -> str:
    """Daily Received vs Sent trend - 2-series line chart with a crosshair-
    free but still legible legend (2 series always gets a legend per the
    skill's rule), gridlines at clean y-ticks, direct end-labels on both
    series so the story is readable without hovering."""
    if not daily:
        return '<p class="chart-empty">No activity in this range yet.</p>'

    pad_l, pad_r, pad_t, pad_b = 36, 54, 16, 28
    plot_w, plot_h = width - pad_l - pad_r, height - pad_t - pad_b

    received = [d["received"] for d in daily]
    sent = [d["sent"] for d in daily]
    hi = max(max(received, default=0), max(sent, default=0), 1)
    # Round the top gridline up to a clean step (per skill: "round to clean
    # numbers"), so ticks read as 0 / 5 / 10 rather than an arbitrary max.
    step = 1
    while step * 5 < hi:
        step *= 2 if step < 10 else 5
    y_max = step * 5

    n = len(daily)
    x_of = lambda i: pad_l + (i / max(n - 1, 1)) * plot_w
    y_of = lambda v: pad_t + plot_h - (v / y_max) * plot_h

    gridlines, y_labels = [], []
    for i in range(6):
        y = pad_t + plot_h - (i / 5) * plot_h
        gridlines.append(f'<line x1="{pad_l}" y1="{y:.1f}" x2="{pad_l + plot_w}" y2="{y:.1f}" stroke="{_GRID}" stroke-width="1"/>')
        y_labels.append(f'<text x="{pad_l - 8}" y="{y + 4:.1f}" text-anchor="end" class="chart-tick">{int(i * step)}</text>')

    def _path(values, color):
        pts = [(x_of(i), y_of(v)) for i, v in enumerate(values)]
        d = "M " + " L ".join(f"{x:.1f},{y:.1f}" for x, y in pts)
        lx, ly = pts[-1]
        dot = f'<circle cx="{lx:.1f}" cy="{ly:.1f}" r="4" fill="{color}" stroke="#fff" stroke-width="2"/>'
        label = f'<text x="{lx + 8:.1f}" y="{ly + 4:.1f}" class="chart-end-label" fill="{color}">{values[-1]}</text>'
        return f'<path d="{d}" fill="none" stroke="{color}" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>{dot}{label}'

    # X-axis labels: first, middle, last day only (per skill: label
    # selectively, never every point) - avoids overlapping text on a wide
    # date range.
    x_label_idxs = sorted(set([0, n // 2, n - 1]))
    x_labels = "".join(
        f'<text x="{x_of(i):.1f}" y="{height - 6}" text-anchor="middle" class="chart-tick">{daily[i]["day"][5:]}</text>'
        for i in x_label_idxs
    )

    return f"""
    <div class="chart-legend">
        <span class="legend-item"><span class="legend-dot" style="background:{_LINE_COLOR_RECEIVED}"></span>Received</span>
        <span class="legend-item"><span class="legend-dot" style="background:{_LINE_COLOR_SENT}"></span>Sent</span>
    </div>
    <svg viewBox="0 0 {width} {height}" width="100%" height="{height}" class="line-chart" preserveAspectRatio="xMidYMid meet">
        {"".join(gridlines)}
        {"".join(y_labels)}
        {_path(received, _LINE_COLOR_RECEIVED)}
        {_path(sent, _LINE_COLOR_SENT)}
        {x_labels}
    </svg>"""


def _bar_chart_svg(rows: list, width: int = 560, height: int = 240) -> str:
    """Top sales reps by lead count - horizontal bars (reads better than
    vertical columns for rep-name labels, which are long and many). Single
    series (lead count) so no legend needed - the panel title already says
    what's plotted. Capped to the top 8 reps so labels stay legible; ties
    resolved by original order."""
    rows = sorted(rows, key=lambda r: r["lead_count"], reverse=True)[:8]
    if not rows:
        return '<p class="chart-empty">No leads in this range yet.</p>'

    max_count = max((r["lead_count"] for r in rows), default=1) or 1
    row_h = 28
    label_w = 130
    bar_area_w = width - label_w - 50
    total_h = len(rows) * row_h + 10

    bars = []
    for i, r in enumerate(rows):
        y = i * row_h + 6
        bar_w = max((r["lead_count"] / max_count) * bar_area_w, 3)
        name = r["rep_name"] if len(r["rep_name"]) <= 18 else r["rep_name"][:16] + "…"
        bars.append(f"""
        <text x="{label_w - 8}" y="{y + 15}" text-anchor="end" class="chart-bar-label">{name}</text>
        <rect x="{label_w}" y="{y}" width="{bar_w:.1f}" height="18" rx="4" fill="{_BRAND_RED}"/>
        <text x="{label_w + bar_w + 8:.1f}" y="{y + 14}" class="chart-bar-value">{r['lead_count']}</text>
        """)

    return f"""<svg viewBox="0 0 {width} {total_h}" width="100%" height="{total_h}" class="bar-chart">
        {"".join(bars)}
    </svg>"""


def _donut_chart_svg(breakdown: list, size: int = 180) -> str:
    """Lead outcome breakdown (New/Contacted/Quoted/Won/Lost) as a donut -
    appropriate here per the skill (part-to-whole, <=6 segments, not a
    close-value comparison). Always shows a legend with the exact count
    per stage (the "relief rule" for the two lower-contrast slices, and
    generally more useful than reading arc angles by eye)."""
    total = sum(b["count"] for b in breakdown)
    r, stroke = size / 2 - 14, 20
    cx = cy = size / 2
    circumference = 2 * 3.14159265 * r

    if total == 0:
        arcs = f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="{_GRID}" stroke-width="{stroke}"/>'
    else:
        offset = 0
        segs = []
        for b in breakdown:
            if b["count"] == 0:
                continue
            frac = b["count"] / total
            dash = frac * circumference
            # 2px surface gap between touching segments (per skill spacer
            # rule) - shorten the dash slightly and let the gap show through.
            gap = 2
            segs.append(
                f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="{_OUTCOME_COLORS[b["outcome"]]}" '
                f'stroke-width="{stroke}" stroke-dasharray="{max(dash - gap, 0):.2f} {circumference - dash + gap:.2f}" '
                f'stroke-dashoffset="{-offset:.2f}" transform="rotate(-90 {cx} {cy})"/>'
            )
            offset += dash
        arcs = "".join(segs)

    legend = "".join(
        f"""<div class="legend-item"><span class="legend-dot" style="background:{_OUTCOME_COLORS[b['outcome']]}"></span>
            {_OUTCOME_PILL_LABELS[b['outcome']]} <span class="chart-legend-count">{b['count']}</span></div>"""
        for b in breakdown
    )

    return f"""
    <div class="donut-wrap">
        <svg viewBox="0 0 {size} {size}" width="{size}" height="{size}" class="donut-chart">
            {arcs}
            <text x="{cx}" y="{cy - 4}" text-anchor="middle" class="donut-total">{total}</text>
            <text x="{cx}" y="{cy + 14}" text-anchor="middle" class="donut-total-label">leads</text>
        </svg>
        <div class="chart-legend chart-legend-vertical">{legend}</div>
    </div>"""


def _render_dashboard_html(start, end, stats, daily, outcome_breakdown, customers, customers_total, customers_page,
                            leads_summary, leads_summary_total, leads_summary_page, leads_summary_all,
                            leads_list, leads_list_total, leads_list_page,
                            rep_replies, reps,
                            selected_phone, transcript, transcript_total, transcript_page,
                            selected_rep_phone, rep_transcript):
    daily_rows = "".join(
        f"<tr><td>{d['day']}</td><td>{d['received']}</td><td>{d['sent']}</td></tr>" for d in daily
    ) or "<tr><td colspan='3' class='muted'>No data in this range</td></tr>"

    total_leads = leads_list_total

    received_sparkline = _sparkline_svg([d["received"] for d in daily], _LINE_COLOR_RECEIVED)
    sent_sparkline = _sparkline_svg([d["sent"] for d in daily], _LINE_COLOR_SENT)
    won_count = next((b["count"] for b in outcome_breakdown if b["outcome"] == "won"), 0)

    line_chart_html = _line_chart_svg(daily)
    donut_chart_html = _donut_chart_svg(outcome_breakdown)
    bar_chart_html = _bar_chart_svg(leads_summary_all)

    leads_summary_rows = "".join(
        f"""<tr>
            <td>{_esc(r['rep_name'])}</td>
            <td>{r['lead_count']}</td>
            <td>{r['customer_count']}</td>
            <td>{_fmt_ts(r['last_lead_at'])}</td>
            <td>{r['failed_notifications'] or '<span class="muted">0</span>'}</td>
        </tr>""" for r in leads_summary
    ) or "<tr><td colspan='5' class='muted'>No leads in this range</td></tr>"

    def _false_positive_action_cell(l) -> str:
        if l.get("false_positive"):
            return _false_positive_pill()
        return f"""<form method="post" action="/dashboard/leads/{l['lead_id']}/mark-false-positive" \
onsubmit="return confirm('Mark this as NOT a real lead? This helps train the Head of WhatsApp Replies.');">
            <input type="hidden" name="phone" value="{_esc(l['phone'])}">
            <input type="hidden" name="enquiry_text" value="{_esc(l['enquiry_text'])}">
            <input type="hidden" name="start" value="{start}">
            <input type="hidden" name="end" value="{end}">
            <button type="submit" class="btn-fp">Mark as false positive</button>
        </form>"""

    def _outcome_cell(l) -> str:
        current = l.get("outcome") or "new"
        options = "".join(
            f'<option value="{v}"{" selected" if v == current else ""}>{label}</option>'
            for v, label in _OUTCOME_PILL_LABELS.items()
        )
        return f"""{_outcome_pill(current, l.get('outcome_amount'))}
        <form method="post" action="/dashboard/leads/{l['lead_id']}/set-outcome" style="margin-top:4px; display:flex; gap:4px;">
            <input type="hidden" name="start" value="{start}">
            <input type="hidden" name="end" value="{end}">
            <input type="hidden" name="lpage" value="{leads_list_page}">
            <select name="outcome" style="font-size:0.78em;">{options}</select>
            <input type="text" name="amount" placeholder="AED (if won)" style="width:90px; font-size:0.78em;">
            <button type="submit" class="btn-fp">Update</button>
        </form>"""

    leads_list_rows = "".join(
        f"""<tr{' style="opacity:0.55"' if l.get('false_positive') else ''}>
            <td>{_fmt_ts(l['created_at'])}</td>
            <td>{_esc(l['company_name']) or _esc(l['phone'])}</td>
            <td>{_esc(l['rep_name'])}</td>
            <td>{_esc(l['enquiry_text'])}</td>
            <td>{_priority_pill(l.get('priority'))}</td>
            <td>{_status_pill(l['status'])}</td>
            <td>{_delivery_pill(l['delivery_status'], l.get('attempt_summary') or '')}</td>
            <td>{_rep_reply_cell(l.get('rep_reply_text'), l.get('rep_reply_at'), l.get('rep_reply_method'))}</td>
            <td>{_outcome_cell(l)}</td>
            <td>{_false_positive_action_cell(l)}</td>
        </tr>""" for l in leads_list
    ) or "<tr><td colspan='10' class='muted'>No leads in this range</td></tr>"

    _CONFIDENCE_LABELS = {"context_match": "Confirmed", "fallback_most_recent": "Best guess", "unresolved": "Unresolved"}
    rep_replies_rows = "".join(
        f"""<tr>
            <td>{_fmt_ts(r['created_at'])}</td>
            <td>{_esc(r['rep_name']) or _esc(r['rep_phone'])}</td>
            <td>{_esc(r['company_name']) or _esc(r['customer_phone']) or '<span class="muted">-</span>'}</td>
            <td>{_esc(r['reply_text'])}</td>
            <td>{_esc(_CONFIDENCE_LABELS.get(r['resolution_method'], r['resolution_method']))}</td>
        </tr>""" for r in rep_replies
    ) or "<tr><td colspan='5' class='muted'>No rep replies in this range</td></tr>"

    customer_rows = "".join(
        f"""<tr class="{'active' if c['phone'] == selected_phone else ''}">
            <td><a href="?start={start}&end={end}&phone={c['phone']}">{_esc(c['phone'])}</a></td>
            <td>{_esc(c['company_name']) or '<span class="muted">-</span>'}</td>
            <td>{_esc(c['rep_name']) or '<span class="muted">-</span>'}</td>
            <td>{c['message_count']}</td>
            <td>{_fmt_ts(c['last_message_at'])}</td>
        </tr>""" for c in customers
    ) or "<tr><td colspan='5' class='muted'>No customers in this range</td></tr>"

    customers_total_pages = max((customers_total + _CUSTOMERS_PAGE_SIZE - 1) // _CUSTOMERS_PAGE_SIZE, 1)
    customers_base_params = {"start": start, "end": end}
    if selected_phone:
        customers_base_params["phone"] = selected_phone
    if selected_rep_phone:
        customers_base_params["rep_phone"] = selected_rep_phone
    customers_pagination_html = _render_pagination(customers_base_params, "cpage", customers_page, customers_total_pages)

    leads_summary_total_pages = max((leads_summary_total + _LEADS_SUMMARY_PAGE_SIZE - 1) // _LEADS_SUMMARY_PAGE_SIZE, 1)
    leads_summary_pagination_html = _render_pagination(
        {"start": start, "end": end}, "spage", leads_summary_page, leads_summary_total_pages,
    )

    leads_list_total_pages = max((leads_list_total + _LEADS_LIST_PAGE_SIZE - 1) // _LEADS_LIST_PAGE_SIZE, 1)
    leads_list_pagination_html = _render_pagination(
        {"start": start, "end": end}, "lpage", leads_list_page, leads_list_total_pages,
    )

    rep_rows = "".join(
        f"""<tr class="{'active' if r['rep_phone'] == selected_rep_phone else ''}">
            <td><a href="?start={start}&end={end}&rep_phone={r['rep_phone']}">{_esc(r['rep_phone'])}</a></td>
            <td>{_esc(r['rep_name']) or '<span class="muted">-</span>'}</td>
            <td>{r['message_count']}</td>
            <td>{_fmt_ts(r['last_activity_at'])}</td>
        </tr>""" for r in reps
    ) or "<tr><td colspan='4' class='muted'>No rep escalations in this range</td></tr>"

    def _manual_escalate_action(m) -> str:
        if m['direction'] != 'in' or m['escalated']:
            return ""
        return f"""<form method="post" action="/dashboard/conversations/{m['id']}/manual-escalate" \
style="margin-top:4px" onsubmit="return confirm('Manually escalate this message to the rep/ops team?');">
            <input type="hidden" name="start" value="{start}">
            <input type="hidden" name="end" value="{end}">
            <input type="hidden" name="phone" value="{_esc(selected_phone)}">
            <input type="hidden" name="page" value="{transcript_page}">
            <button type="submit" class="btn-fp">Escalate</button>
        </form>"""

    if transcript is not None:
        if transcript:
            bubbles = "".join(
                f"""<div class="bubble {'in' if m['direction'] == 'in' else 'out'} {'escalated' if m['escalated'] else ''}">
                    <div class="bubble-text">{_esc(m['message'])}</div>
                    <div class="bubble-time">{_fmt_ts(m['created_at'])}{' &middot; escalated' if m['escalated'] else ''}</div>
                    {_manual_escalate_action(m)}
                </div>""" for m in transcript
            )
        else:
            bubbles = "<p class='muted'>No messages for this customer in the selected date range.</p>"
        total_pages = max((transcript_total + _TRANSCRIPT_PAGE_SIZE - 1) // _TRANSCRIPT_PAGE_SIZE, 1)
        pagination_html = _render_pagination(
            {"start": start, "end": end, "phone": selected_phone}, "page", transcript_page, total_pages,
        )
        transcript_html = f"""
        <div class="panel">
            <h2>Transcript &middot; {_esc(selected_phone)} <span class="muted" style="font-weight:normal">({transcript_total} messages)</span></h2>
            <div class="chat-window">{bubbles}</div>
            {pagination_html}
        </div>"""
    else:
        transcript_html = """
        <div class="panel">
            <h2>Transcript</h2>
            <p class="muted">Select a customer from the list to view their conversation.</p>
        </div>"""

    if rep_transcript is not None:
        if rep_transcript:
            rep_bubbles = "".join(
                f"""<div class="bubble {'out' if m['direction'] == 'out' else 'in'}">
                    <div class="bubble-text">{(_esc(m['company_name']) + ' &mdash; ') if m['company_name'] else ''}{_esc(m['message_text'])}</div>
                    <div class="bubble-time">{_fmt_ts(m['created_at'])}{' &middot; ' + _esc(_CONFIDENCE_LABELS.get(m['extra'], m['extra'])) if m['direction'] == 'in' and m.get('extra') else ''}</div>
                </div>""" for m in rep_transcript
            )
        else:
            rep_bubbles = "<p class='muted'>No escalation activity for this rep in the selected date range.</p>"
        rep_transcript_html = f"""
        <div class="panel">
            <h2>Rep escalation transcript &middot; {_esc(selected_rep_phone)}</h2>
            <div class="chat-window">{rep_bubbles}</div>
        </div>"""
    else:
        rep_transcript_html = """
        <div class="panel">
            <h2>Rep escalation transcript</h2>
            <p class="muted">Select a rep from the list to view their escalation alerts and replies.</p>
        </div>"""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Würth WhatsApp Agent - Dashboard</title>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta name="robots" content="noindex, nofollow">
<style>
{_BASE_STYLE}
  .container {{ padding: 24px; max-width: 1320px; margin: 0 auto; }}
  .filters {{ background: white; border: 1px solid #eceef1; border-radius: 12px; padding: 14px 18px; margin-bottom: 20px;
              display: flex; gap: 14px; align-items: center; flex-wrap: wrap; }}
  .filters label {{ font-size: 0.82em; color: #6b7280; font-weight: 600; display: flex; align-items: center; gap: 6px; }}
  .filters input {{ padding: 7px 10px; border: 1px solid #e2e4e9; border-radius: 8px; width: 100%; max-width: 160px; font-size: 0.9em; }}
  .filters input:focus {{ outline: none; border-color: #CC0000; box-shadow: 0 0 0 3px rgba(204,0,0,0.1); }}
  .filters button, .filters a.button {{ background: #CC0000; color: white; border: none; padding: 9px 18px; border-radius: 8px;
              cursor: pointer; text-decoration: none; font-size: 0.88em; font-weight: 600; display: inline-flex; align-items: center; gap: 6px; }}
  .filters button:hover, .filters a.button:hover {{ background: #a80000; }}
  .filters a.button.secondary {{ background: white; color: #1a1a1a; border: 1px solid #e2e4e9; }}
  .filters a.button.secondary:hover {{ background: #f7f7f9; }}

  .stats {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; margin-bottom: 20px; }}
  .stat-card {{ background: white; border: 1px solid #eceef1; border-radius: 14px; padding: 18px 20px; display: flex;
                flex-direction: column; gap: 2px; position: relative; overflow: hidden; }}
  .stat-card .stat-top {{ display: flex; align-items: flex-start; justify-content: space-between; gap: 8px; }}
  .stat-card .icon {{ width: 34px; height: 34px; border-radius: 9px; display: flex; align-items: center; justify-content: center;
                       font-size: 1.05em; flex-shrink: 0; }}
  .stat-card .icon.i-messages {{ background: #eaf1fd; }}
  .stat-card .icon.i-replies {{ background: #fdeaea; }}
  .stat-card .icon.i-customers {{ background: #eafaf1; }}
  .stat-card .icon.i-leads {{ background: #fff6e6; }}
  .stat-card .icon.i-won {{ background: #e9f7ec; }}
  .stat-card .value {{ font-size: 1.7em; font-weight: 700; color: #1a1a1a; margin-top: 10px; letter-spacing: -0.02em; }}
  .stat-card .label {{ font-size: 0.82em; color: #6b7280; font-weight: 600; }}
  .stat-card .sparkline {{ position: absolute; right: 16px; bottom: 14px; opacity: 0.9; }}

  .charts-grid {{ display: grid; grid-template-columns: 1.4fr 1fr; gap: 16px; margin-bottom: 20px; align-items: stretch; }}
  @media (max-width: 980px) {{ .charts-grid {{ grid-template-columns: 1fr; }} }}
  .chart-row {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 20px; }}
  @media (max-width: 980px) {{ .chart-row {{ grid-template-columns: 1fr; }} }}

  .grid {{ display: grid; grid-template-columns: 1.2fr 1fr; gap: 16px; align-items: start; }}
  @media (max-width: 900px) {{ .grid {{ grid-template-columns: 1fr; }} }}

  .panel {{ background: white; border: 1px solid #eceef1; border-radius: 14px; padding: 18px 20px; margin-bottom: 20px; overflow-x: auto; }}
  .panel h2 {{ font-size: 0.98em; margin: 0 0 4px 0; font-weight: 700; display: flex; align-items: center; gap: 8px; }}
  .panel h2 .badge {{ background: #f7f7f9; color: #6b7280; font-size: 0.72em; font-weight: 700; padding: 2px 8px; border-radius: 20px; }}
  .panel .subtitle {{ font-size: 0.82em; color: #6b7280; margin: 0 0 14px 0; }}

  table {{ width: 100%; border-collapse: collapse; font-size: 0.85em; min-width: 380px; }}
  th, td {{ text-align: left; padding: 9px 10px; border-bottom: 1px solid #f2f3f5; white-space: nowrap; }}
  th {{ color: #9ca3af; font-weight: 700; font-size: 0.72em; text-transform: uppercase; letter-spacing: 0.03em; }}
  tr.active {{ background: #fdf1f1; }}
  tr:hover {{ background: #fafafb; }}
  .muted {{ color: #9ca3af; }}

  .pill {{ display: inline-block; padding: 3px 10px; border-radius: 20px; font-size: 0.76em; font-weight: 700; white-space: nowrap; }}
  .pill.delivered {{ background: #e6f4ea; color: #1a7f37; }}
  .pill.failed {{ background: #fdeaec; color: #CC0000; }}
  .pill.pending {{ background: #f0f0f0; color: #888; }}
  .pill.status-open {{ background: #e8f0fe; color: #1a56c8; }}
  .pill.status-closed {{ background: #f0f0f0; color: #888; }}
  .pill.guess {{ background: #fff4e5; color: #a85d00; }}
  .pill.priority-high {{ background: #fdeaec; color: #CC0000; }}
  .pill.priority-medium {{ background: #fff4e5; color: #a85d00; }}
  .pill.priority-low {{ background: #f0f0f0; color: #888; }}
  .pill.fp {{ background: #f0f0f0; color: #888; }}
  .pill.outcome-new {{ background: #eaf1fd; color: #2a78d6; }}
  .pill.outcome-contacted {{ background: #e4f7ee; color: #1baf7a; }}
  .pill.outcome-quoted {{ background: #fff6e6; color: #a86e00; }}
  .pill.outcome-won {{ background: #e6f4ea; color: #008300; }}
  .pill.outcome-lost {{ background: #fdeaea; color: #e34948; }}

  .btn-fp {{ background: white; border: 1px solid #e2e4e9; color: #4b5563; font-size: 0.78em; font-weight: 600; padding: 4px 10px;
             border-radius: 7px; cursor: pointer; white-space: nowrap; }}
  .btn-fp:hover {{ background: #f7f7f9; border-color: #CC0000; color: #CC0000; }}

  .chat-window {{ max-height: 500px; overflow-y: auto; display: flex; flex-direction: column; gap: 8px; }}
  .bubble {{ max-width: 85%; padding: 9px 13px; border-radius: 14px; font-size: 0.9em; }}
  .bubble.in {{ align-self: flex-start; background: #f2f3f5; border-bottom-left-radius: 4px; }}
  .bubble.out {{ align-self: flex-end; background: #fdecec; border-bottom-right-radius: 4px; }}
  .bubble.escalated {{ border: 1px solid #CC0000; }}
  .bubble-text {{ white-space: pre-wrap; word-break: break-word; }}
  .bubble-time {{ font-size: 0.7em; color: #9ca3af; margin-top: 4px; }}

  .pagination {{ display: flex; gap: 6px; flex-wrap: wrap; margin-top: 14px; padding-top: 14px; border-top: 1px solid #f2f3f5; }}
  .page-num {{ display: inline-block; min-width: 28px; text-align: center; padding: 5px 9px; border-radius: 7px; font-size: 0.85em;
               text-decoration: none; color: #4b5563; background: #f7f7f9; font-weight: 600; }}
  .page-num:hover {{ background: #eceef1; }}
  .page-num.active {{ background: #CC0000; color: white; }}

  /* ===== Charts ===== */
  .chart-empty {{ color: #9ca3af; font-size: 0.85em; padding: 40px 0; text-align: center; }}
  .chart-legend {{ display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 10px; font-size: 0.8em; color: #4b5563; font-weight: 600; }}
  .chart-legend-vertical {{ flex-direction: column; gap: 8px; margin-bottom: 0; align-items: flex-start; }}
  .legend-item {{ display: inline-flex; align-items: center; gap: 6px; }}
  .legend-dot {{ width: 9px; height: 9px; border-radius: 50%; display: inline-block; flex-shrink: 0; }}
  .chart-legend-count {{ color: #9ca3af; font-weight: 700; }}
  .chart-tick {{ font-size: 9px; fill: #9ca3af; font-weight: 600; }}
  .chart-end-label {{ font-size: 11px; font-weight: 700; }}
  .chart-bar-label {{ font-size: 11px; fill: #4b5563; font-weight: 600; }}
  .chart-bar-value {{ font-size: 11px; fill: #1a1a1a; font-weight: 700; dominant-baseline: middle; }}
  .donut-wrap {{ display: flex; align-items: center; gap: 20px; flex-wrap: wrap; justify-content: center; }}
  .donut-total {{ font-size: 26px; font-weight: 700; fill: #1a1a1a; }}
  .donut-total-label {{ font-size: 10px; fill: #9ca3af; font-weight: 600; text-transform: uppercase; letter-spacing: 0.05em; }}
  .line-chart, .bar-chart {{ display: block; }}

  @media (max-width: 600px) {{
    .container {{ padding: 14px; }}
    header h1 {{ font-size: 1em; }}
    .filters {{ flex-direction: column; align-items: stretch; }}
    .filters input {{ max-width: none; }}
    .stats {{ grid-template-columns: repeat(2, 1fr); }}
  }}
</style>
</head>
<body>
<header>
  <img class="logo" src="{LOGO_URL}" alt="Würth">
  <h1>WhatsApp Agent &middot; Sales Dashboard</h1>
  <a class="logout" href="/dashboard/logout">Log out &rarr;</a>
</header>
<div class="container">

  <form class="filters" method="get">
    <label>📅 From <input type="date" name="start" value="{start}"></label>
    <label>To <input type="date" name="end" value="{end}"></label>
    <button type="submit">Apply</button>
    <a class="button secondary" href="/dashboard/export?start={start}&end={end}">⬇ Export to Excel</a>
  </form>

  <div class="stats">
    <div class="stat-card">
      <div class="stat-top"><span class="icon i-messages">💬</span></div>
      <div class="value">{stats['messages_received']}</div>
      <div class="label">Messages received</div>
      {received_sparkline}
    </div>
    <div class="stat-card">
      <div class="stat-top"><span class="icon i-replies">↩️</span></div>
      <div class="value">{stats['replies_sent']}</div>
      <div class="label">Replies sent</div>
      {sent_sparkline}
    </div>
    <div class="stat-card">
      <div class="stat-top"><span class="icon i-customers">👥</span></div>
      <div class="value">{stats['unique_customers']}</div>
      <div class="label">Unique customers</div>
    </div>
    <div class="stat-card">
      <div class="stat-top"><span class="icon i-leads">🎯</span></div>
      <div class="value">{total_leads}</div>
      <div class="label">Sales leads generated</div>
    </div>
    <div class="stat-card">
      <div class="stat-top"><span class="icon i-won">🏆</span></div>
      <div class="value">{won_count}</div>
      <div class="label">Deals won</div>
    </div>
  </div>

  <div class="charts-grid">
    <div class="panel">
      <h2>Message activity</h2>
      <p class="subtitle">Customer messages received vs. AI replies sent, per day</p>
      {line_chart_html}
    </div>
    <div class="panel">
      <h2>Lead pipeline</h2>
      <p class="subtitle">Every open &amp; closed lead by outcome stage</p>
      {donut_chart_html}
    </div>
  </div>

  <div class="panel">
    <h2>Top sales reps by leads <span class="badge">{leads_summary_total} reps</span></h2>
    <p class="subtitle">How AI-detected leads are distributed across the team in this range</p>
    {bar_chart_html}
  </div>

  <div class="panel">
    <h2>Daily activity</h2>
    <table>
      <tr><th>Date</th><th>Received</th><th>Sent</th></tr>
      {daily_rows}
    </table>
  </div>

  <div class="panel">
    <h2>Sales leads by rep &middot; how AI is helping the team ({leads_summary_total} reps, {total_leads} leads total)</h2>
    <p class="subtitle">A "lead" is a customer enquiry the AI recognized as purchase intent, a quote/pricing \
request, or an urgent issue, and flagged for the assigned sales rep to follow up on.</p>
    <table>
      <tr><th>Sales Rep</th><th>Leads</th><th>Customers</th><th>Last lead</th><th>Failed notifications</th></tr>
      {leads_summary_rows}
    </table>
    {leads_summary_pagination_html}
  </div>

  <div class="panel">
    <h2>Recent leads <span class="badge">{leads_list_total}</span></h2>
    <table>
      <tr><th>When</th><th>Customer</th><th>Rep</th><th>Enquiry</th><th>Priority</th><th>Status</th><th>Delivery</th><th>Rep Response</th><th>Outcome</th><th>Action</th></tr>
      {leads_list_rows}
    </table>
    {leads_list_pagination_html}
  </div>

  <div class="panel">
    <h2>Rep replies <span class="badge">{len(rep_replies)}</span></h2>
    <p class="subtitle">Every reply a sales rep sent back after being notified of a lead, in order - a dedicated \
view of rep engagement separate from the leads table above (which only shows the latest reply per lead).</p>
    <table>
      <tr><th>When</th><th>Rep</th><th>Customer</th><th>Reply</th><th>Match confidence</th></tr>
      {rep_replies_rows}
    </table>
  </div>

  <div class="grid">
    <div class="panel">
      <h2>Customers <span class="badge">{customers_total}</span></h2>
      <table>
        <tr><th>Phone</th><th>Company</th><th>Rep</th><th>Msgs</th><th>Last active</th></tr>
        {customer_rows}
      </table>
      {customers_pagination_html}
    </div>
    {transcript_html}
  </div>

  <div class="grid">
    <div class="panel">
      <h2>Sales reps <span class="badge">{len(reps)}</span></h2>
      <table>
        <tr><th>Phone</th><th>Rep</th><th>Msgs</th><th>Last active</th></tr>
        {rep_rows}
      </table>
    </div>
    {rep_transcript_html}
  </div>

</div>
</body>
</html>"""
